import smtplib
import pandas as pd
import requests
import sys
import datetime
from bs4 import BeautifulSoup
from price_parser import Price
from openpyxl import load_workbook

PRODUCT_URL_FILE = "products.xlsx"
PRICES_FILE = "prices.xlsx"
SAVE_TO_FILE = True
SEND_MAIL = False


def main():
    df = get_urls(PRODUCT_URL_FILE)
    df_updated = process_products(df)
    if SAVE_TO_FILE:
        with pd.ExcelWriter(PRICES_FILE, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
            wb = load_workbook(PRICES_FILE, read_only=True)
            rownum = writer.sheets["Sheet1"].max_row
            df_updated.to_excel(writer,index=False,header=False,startrow=rownum)
    if SEND_MAIL:
        send_mail(df_updated)


def get_urls(file):
    df = pd.read_excel(file)
    return df


def process_products(df):
    update_products = []
    for product in df.to_dict("records"):
        html = product["url"]
        product["price"] = get_price(html)
        product["date_time"] = datetime.datetime.now()
        update_products.append(product)
    return pd.DataFrame(update_products)


def get_price(html):
    soup = BeautifulSoup(extract_source(html), "lxml")
    if soup.find_all(class_="sale-price"):
        el = soup.select_one(".sale-price")
    elif soup.find_all(class_="sales-price"):
        el = soup.select_one(".sales-price")
    else:
        print(f"Terminated early on {html}")
        sys.exit()
    price = Price.fromstring(el.text)
    return price.amount_float


def extract_source(url):
    agent = {"User-Agent": "Mozilla/5.0"}
    source = requests.get(url, headers=agent).text
    return source


def send_mail(df):
    message_text = get_mail(df)
    with smtplib.SMTP("smtp address", port) as smtp:
        smtp.starttls()
        smtp.login("email address", "password")
        smtp.sendmail("email address",
                      "email address", message_text)


def get_mail(df):
    subject = "Price Drop Alert"
    body = df["alert"].to_string()
    subject_and_message = f"Subject:{subject}\n\n{body}"
    return subject_and_message


if __name__ == "__main__":
    main()
